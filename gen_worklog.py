#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成工作量登记表 Excel 文件"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date
import calendar

wb = Workbook()

# ===== 颜色定义 =====
C_HEADER_BG  = "1F4E79"
C_SUBHEADER  = "2E75B6"
C_ACCENT     = "4472C4"
C_LIGHT_BLUE = "D6E4F0"
C_WHITE      = "FFFFFF"
C_DARK_TEXT  = "1F1F1F"
C_GREEN      = "375623"
C_GREEN_BG   = "E2EFDA"
C_YELLOW_BG  = "FFF2CC"
C_ORANGE_BG  = "FCE4D6"
C_GRAY_BG    = "F2F2F2"


def thin_border():
    side = Side(style="thin", color="BFBFBF")
    return Border(left=side, right=side, top=side, bottom=side)


def set_cell(ws, row, col, value, bg=None, bold=False,
             font_color=None, align="center", fg_color=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(
        name="微软雅黑",
        bold=bold,
        color=font_color or C_DARK_TEXT,
        size=10 if not bold else 11
    )
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = thin_border()
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    return cell


# ============================================================
#  Sheet 1: 工作日志
# ============================================================
ws_log = wb.active
ws_log.title = "📋 工作日志"
ws_log.sheet_view.showGridLines = False

# --- 标题 ---
ws_log.merge_cells("A1:J1")
title = ws_log.cell(1, 1, "工作量登记表")
title.font = Font(name="微软雅黑", bold=True, size=16, color=C_WHITE)
title.fill = PatternFill("solid", fgColor=C_HEADER_BG)
title.alignment = Alignment(horizontal="center", vertical="center")
ws_log.row_dimensions[1].height = 38

ws_log.merge_cells("A2:J2")
sub = ws_log.cell(2, 1, f"统计周期：{date.today().year}年  填写人：money")
sub.font = Font(name="微软雅黑", size=10, color="666666")
sub.alignment = Alignment(horizontal="center", vertical="center")
ws_log.row_dimensions[2].height = 20

# --- 列宽 ---
col_widths = [12, 8, 22, 10, 22, 10, 10, 10, 14, 20]
for i, w in enumerate(col_widths, 1):
    ws_log.column_dimensions[get_column_letter(i)].width = w

# --- 表头 ---
headers = [
    "日期", "星期", "上午工作内容", "上午耗时\n(小时)",
    "下午工作内容", "下午耗时\n(小时)", "加班耗时\n(小时)",
    "当日合计\n(小时)", "工作类型", "备注"
]
row_header = 4
for i, h in enumerate(headers, 1):
    set_cell(ws_log, row_header, i, h,
             bg=C_SUBHEADER, bold=True, font_color=C_WHITE)
ws_log.row_dimensions[row_header].height = 36

# --- 数据验证（工作类型）---
type_options = ",".join([
    "功能开发", "Bug修复", "代码审查", "需求分析",
    "文档编写", "测试验证", "会议沟通", "技术调研",
    "运维部署", "培训分享", "其他"
])
dv_type = DataValidation(
    type="list",
    formula1=f'"{type_options}"',
    allow_blank=True
)
dv_type.prompt = "请选择工作类型"
dv_type.promptTitle = "工作类型"
ws_log.add_data_validation(dv_type)

# --- 预填当月日期 ---
today = date.today()
year, month = today.year, today.month
_, days_in_month = calendar.monthrange(year, month)

row_start = row_header + 1
weekday_names = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]

for d in range(1, days_in_month + 1):
    row = row_start + d - 1
    current = date(year, month, d)
    wd = weekday_names[current.weekday()]
    bg = C_LIGHT_BLUE if d % 2 == 0 else C_WHITE
    is_weekend = current.weekday() >= 5
    wd_bg = C_ORANGE_BG if is_weekend else bg

    set_cell(ws_log, row, 1, current.strftime("%Y-%m-%d"), bg=bg, align="center")
    set_cell(ws_log, row, 2, wd, bg=wd_bg, align="center")

    # 上午工作 / 上午耗时 / 下午工作 / 下午耗时 / 加班耗时
    for c in [3, 4, 5, 6, 7]:
        set_cell(ws_log, row, c, None, bg=bg, align="left")

    # 当日合计公式
    cell = ws_log.cell(row=row, column=8, value=f"=SUM(D{row},F{row},G{row})")
    cell.font = Font(name="微软雅黑", bold=True, size=10, color=C_DARK_TEXT)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()
    cell.fill = PatternFill("solid", fgColor=C_YELLOW_BG)

    # 工作类型下拉
    cell = ws_log.cell(row=row, column=9)
    cell.font = Font(name="微软雅黑", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()
    cell.fill = PatternFill("solid", fgColor=bg)
    dv_type.add(cell)

    # 备注
    set_cell(ws_log, row, 10, None, bg=bg, align="left")

    ws_log.row_dimensions[row].height = 30

# 冻结首4行
ws_log.freeze_panes = "A5"

# ============================================================
#  Sheet 2: 月度汇总
# ============================================================
ws_sum = wb.create_sheet("📊 月度汇总")
ws_sum.sheet_view.showGridLines = False
ws_sum.column_dimensions["A"].width = 18
ws_sum.column_dimensions["B"].width = 12
ws_sum.column_dimensions["C"].width = 12
ws_sum.column_dimensions["D"].width = 14
ws_sum.column_dimensions["E"].width = 12
ws_sum.column_dimensions["F"].width = 22

# 标题
ws_sum.merge_cells("A1:F1")
t = ws_sum.cell(1, 1, f"{year}年{month}月 工作量汇总")
t.font = Font(name="微软雅黑", bold=True, size=14, color=C_WHITE)
t.fill = PatternFill("solid", fgColor=C_HEADER_BG)
t.alignment = Alignment(horizontal="center", vertical="center")
ws_sum.row_dimensions[1].height = 36

# 表头
sum_headers = ["工作类型", "本月工时", "占比(%)", "上月对比", "趋势", "述职要点"]
for i, h in enumerate(sum_headers, 1):
    set_cell(ws_sum, 3, i, h, bg=C_SUBHEADER, bold=True, font_color=C_WHITE)
ws_sum.row_dimensions[3].height = 30

type_rows = [
    "功能开发", "Bug修复", "代码审查", "需求分析",
    "文档编写", "测试验证", "会议沟通", "技术调研",
    "运维部署", "培训分享", "其他"
]

for i, tname in enumerate(type_rows, 4):
    ws_sum.row_dimensions[i].height = 26
    set_cell(ws_sum, i, 1, tname, bg=C_LIGHT_BLUE)
    # 本月工时（用 COUNTIF + SUMIF 逻辑手动填写提示）
    h2 = ws_sum.cell(row=i, column=2, value=0)
    h2.font = Font(name="微软雅黑", size=10, bold=True)
    h2.alignment = Alignment(horizontal="center", vertical="center")
    h2.border = thin_border()
    h2.fill = PatternFill("solid", fgColor=C_GREEN_BG)
    # 占比
    ws_sum.cell(row=i, column=3).value = f"=IF($B$16>0,B{i}/$B$16*100,\"\")"
    set_cell(ws_sum, i, 3, ws_sum.cell(row=i, column=3).value,
             bg=C_GRAY_BG, align="center")
    set_cell(ws_sum, i, 4, "待填写", bg=C_GRAY_BG, align="center")
    set_cell(ws_sum, i, 5, "→", bg=C_GRAY_BG, align="center")
    set_cell(ws_sum, i, 6, "填写核心成果关键词", bg=C_GRAY_BG, align="left")

# 合计行
row_total = 4 + len(type_rows)
ws_sum.row_dimensions[row_total].height = 28
set_cell(ws_sum, row_total, 1, "合  计", bg=C_GREEN, bold=True, font_color=C_WHITE)
ws_sum.cell(row=row_total, column=2).value = f"=SUM(B4:B{row_total-1})"
h = ws_sum.cell(row=row_total, column=2)
h.font = Font(name="微软雅黑", bold=True, size=11, color=C_DARK_TEXT)
h.alignment = Alignment(horizontal="center", vertical="center")
h.border = thin_border()
h.fill = PatternFill("solid", fgColor=C_GREEN_BG)
set_cell(ws_sum, row_total, 3, "100%", bg=C_GREEN, bold=True, font_color=C_WHITE)
set_cell(ws_sum, row_total, 4, "", bg=C_GREEN_BG)
set_cell(ws_sum, row_total, 5, "", bg=C_GREEN_BG)
set_cell(ws_sum, row_total, 6, "用于述职报告「工作量总览」部分",
         bg=C_GREEN_BG, bold=True)

# ============================================================
#  Sheet 3: 述职素材
# ============================================================
ws_mat = wb.create_sheet("📝 述职素材")
ws_mat.sheet_view.showGridLines = False
ws_mat.merge_cells("A1:D1")
t = ws_mat.cell(1, 1, "述职报告素材库（按季度/年度整理）")
t.font = Font(name="微软雅黑", bold=True, size=14, color=C_WHITE)
t.fill = PatternFill("solid", fgColor=C_HEADER_BG)
t.alignment = Alignment(horizontal="center", vertical="center")
ws_mat.row_dimensions[1].height = 36

ws_mat.column_dimensions["A"].width = 16
ws_mat.column_dimensions["B"].width = 30
ws_mat.column_dimensions["C"].width = 30
ws_mat.column_dimensions["D"].width = 30

mat_headers = ["类别", "核心成果（量化）", "亮点/创新点", "下一步计划"]
for i, h in enumerate(mat_headers, 1):
    set_cell(ws_mat, 3, i, h, bg=C_SUBHEADER, bold=True, font_color=C_WHITE)
ws_mat.row_dimensions[3].height = 30

mat_rows = [
    "重点项目", "技术突破", "团队协作", "流程优化",
    "个人成长", "待改进项"
]
for i, name in enumerate(mat_rows, 4):
    ws_mat.row_dimensions[i].height = 50
    set_cell(ws_mat, i, 1, name, bg=C_ACCENT, bold=True, font_color=C_WHITE)
    for c in [2, 3, 4]:
        set_cell(ws_mat, i, c, "", bg=C_GRAY_BG, align="left")

# ============================================================
#  Sheet 4: 使用说明
# ============================================================
ws_hlp = wb.create_sheet("📖 使用说明")
ws_hlp.sheet_view.showGridLines = False
ws_hlp.merge_cells("A1:C1")
t = ws_hlp.cell(1, 1, "使用说明 — 工作量登记表")
t.font = Font(name="微软雅黑", bold=True, size=14, color=C_WHITE)
t.fill = PatternFill("solid", fgColor=C_HEADER_BG)
t.alignment = Alignment(horizontal="center", vertical="center")
ws_hlp.row_dimensions[1].height = 36

ws_hlp.column_dimensions["A"].width = 22
ws_hlp.column_dimensions["B"].width = 50
ws_hlp.column_dimensions["C"].width = 18

hlp_headers = ["功能", "说明", "述职阶段用途"]
for i, h in enumerate(hlp_headers, 1):
    set_cell(ws_hlp, 3, i, h, bg=C_SUBHEADER, bold=True, font_color=C_WHITE)
ws_hlp.row_dimensions[3].height = 30

tips = [
    ("填写工作日志",
     "每天填写「工作日志」表，工作类型可从下拉框选择，当日合计自动计算。",
     "提供原始素材"),
    ("月度汇总统计",
     "每月底填写「月度汇总」表，统计各类型工时占比，识别工作重点。",
     "述职报告「工作量分布」章节"),
    ("述职素材库",
     "在「述职素材」表中按类别填写核心成果（尽量量化：节省X小时、提升X%）。",
     "述职报告主体内容"),
    ("加班/周末记录",
     "在「加班耗时」列记录，周末行已用橙色标记，方便核对。",
     "体现工作投入度"),
    ("多月份扩展",
     "复制「工作日志」表，重命名为「3月」「4月」等，方便跨月对比。",
     "季度/年度述职对比"),
    ("数据导出",
     "可将「月度汇总」数据直接复制到 PPT 或述职文档中。",
     "制作述职报告"),
]

for i, (title, desc, use) in enumerate(tips, 4):
    ws_hlp.row_dimensions[i].height = 50
    set_cell(ws_hlp, i, 1, title, bg=C_ACCENT, bold=True,
             font_color=C_WHITE, align="left")
    set_cell(ws_hlp, i, 2, desc, bg=C_GRAY_BG, align="left")
    set_cell(ws_hlp, i, 3, use, bg=C_GREEN_BG, align="center", bold=True)

# ============================================================
#  保存
# ============================================================
output_path = "/Users/money/WorkBuddy/20260505213642/工作量登记表.xlsx"
wb.save(output_path)
print(f"✅ 文件已生成：{output_path}")
